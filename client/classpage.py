import tkinter as tk
from  models import *
from tkinter import ttk,filedialog, messagebox
import os
from tkcalendar import DateEntry
from tktimepicker import AnalogPicker,AnalogThemes, constants
from openpyxl import load_workbook
import requests
import datetime
import torch
from core.detection.detec import load_retinaface
from core.recognition.rec import load_ghostfacenets
from core.utils import image_to_vector
from extension import db,app
class ClassPage(tk.Frame):
    def __init__(self, root):

        super().__init__(root)
        self.pack(fill=tk.BOTH, expand=True)
        self.root = root
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=19)
        self.columnconfigure(0, weight=1)

        self.style = ttk.Style(self)
        self.style.theme_use("forest-light")

        self.widget_frame_menu = ttk.LabelFrame(self, text="Menu")
        self.widget_frame_menu.grid(row=0, column=0, padx=10, pady=2, sticky="ew")

        accentbutton = ttk.Button(self.widget_frame_menu, text="Thêm lớp", style="Accent.TButton",command=self.add_class)
        accentbutton.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.list_class = self.find_all_class()
        self.select_class_var = tk.StringVar()
        self.select_class = ttk.OptionMenu(self.widget_frame_menu,self.select_class_var ,  *self.list_class)
        self.select_class.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        self.button_select_class = ttk.Button(self.widget_frame_menu, text="Xem", style="Accent.TButton",command=self.take_class_data)
        self.button_select_class.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")


        # Create a Frame for the Treeview
        self.treeFrame = ttk.Frame(self)
        self.treeFrame.grid(row=1, column=0, padx=10, pady=2, sticky="nsew")

        # Scrollbar
        treeScroll = ttk.Scrollbar(self.treeFrame)
        treeScroll.pack(side="right", fill="y")

        # Treeview
        treeview = ttk.Treeview(self.treeFrame, selectmode="extended", yscrollcommand=treeScroll.set, columns=(1, 2),
                                height=12)
        treeview.pack(expand=True, fill="both")
        treeScroll.config(command=treeview.yview)

        # Treeview columns
        treeview.column("#0", width=50)
        treeview.column(1, anchor="w", width=50)
        treeview.column(2, anchor="w", width=50)

        # Treeview headings
        treeview.heading("#0", text="MSSV", anchor="center")
        treeview.heading(1, text="MSSV", anchor="center")
        treeview.heading(2, text="Họ và tên", anchor="center")

        # Define treeview data
        treeview_data = [
            ("", "end", 1, "Parent", ("Item 1", "Value 1")),
            (1, "end", 2, "Child", ("Subitem 1.1", "Value 1.1")),
            (1, "end", 3, "Child", ("Subitem 1.2", "Value 1.2")),
            (1, "end", 4, "Child", ("Subitem 1.3", "Value 1.3")),
            (1, "end", 5, "Child", ("Subitem 1.4", "Value 1.4")),
            ("", "end", 6, "Parent", ("Item 2", "Value 2")),
            (6, "end", 7, "Child", ("Subitem 2.1", "Value 2.1")),
            (6, "end", 8, "Sub-parent", ("Subitem 2.2", "Value 2.2")),
            (8, "end", 9, "Child", ("Subitem 2.2.1", "Value 2.2.1")),
            (8, "end", 10, "Child", ("Subitem 2.2.2", "Value 2.2.2")),
            (8, "end", 11, "Child", ("Subitem 2.2.3", "Value 2.2.3")),
            (6, "end", 12, "Child", ("Subitem 2.3", "Value 2.3")),
            (6, "end", 13, "Child", ("Subitem 2.4", "Value 2.4")),
            ("", "end", 14, "Parent", ("Item 3", "Value 3")),
            (14, "end", 15, "Child", ("Subitem 3.1", "Value 3.1")),
            (14, "end", 16, "Child", ("Subitem 3.2", "Value 3.2")),
            (14, "end", 17, "Child", ("Subitem 3.3", "Value 3.3")),
            (14, "end", 18, "Child", ("Subitem 3.4", "Value 3.4")),
            ("", "end", 19, "Parent", ("Item 4", "Value 4")),
            (19, "end", 20, "Child", ("Subitem 4.1", "Value 4.1")),
            (19, "end", 21, "Sub-parent", ("Subitem 4.2", "Value 4.2")),
            (21, "end", 22, "Child", ("Subitem 4.2.1", "Value 4.2.1")),
            (21, "end", 23, "Child", ("Subitem 4.2.2", "Value 4.2.2")),
            (21, "end", 24, "Child", ("Subitem 4.2.3", "Value 4.2.3")),
            (19, "end", 25, "Child", ("Subitem 4.3", "Value 4.3"))
        ]

        # Insert treeview data
        for item in treeview_data:
            treeview.insert(parent=item[0], index=item[1], iid=item[2], text=item[3], values=item[4])
            if item[0] == "" or item[2] in (8, 12):
                treeview.item(item[2], open=True)  # Open parents

        # Select and scroll
        treeview.selection_set(10)
        treeview.see(7)
    def take_class_data(self):
        class_name = self.select_class_var.get()
        #select class
        with app.app_context():
            class_ = db.session.query(Class).filter_by(name=class_name).first()
            #select all student in class
            enrollments = db.session.query(Enrollment).filter_by(class_id=class_.id).all()
            #select all session in class
            class_sessions = db.session.query(ClassSession).filter_by(course_id=class_.id).all()
            #select all grade in class
            grades = db.session.query(Grade).filter_by(class_id=class_.id).all()
            #select all attendance in class
            attendance_records = db.session.query(AttendanceRecord).filter_by(class_session_id=class_.id).all()
            print(class_,enrollments,class_sessions,grades,attendance_records)

    def find_all_class(self):
        with app.app_context():
            classes = db.session.query(Class).all()
            #take all name of class to list
            classes = [class_.name for class_ in classes]
            print("class,",classes)
            return classes
    def add_class(self):
        form_add_class = tk.Toplevel(self)

        form_add_class.title("Thêm lớp")
        form_add_class.geometry("800x400")

        form_add_class.columnconfigure(0, weight=1)
        # Biến lưu giá trị nhập vào
        class_name_var = tk.StringVar()
        subject_name_var = tk.StringVar()
        session_count_var = tk.IntVar(value=1)


        period_count_var = tk.IntVar(value=1)
        file_path_var = tk.StringVar()

        # Hàm chọn file Excel
        def select_file():
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx")],  # Chỉ chấp nhận file .xlsx
                title="Chọn tệp Excel"
            )
            if file_path.endswith(".xlsx"):
                file_path_var.set(file_path)
            else:
                messagebox.showerror("Lỗi", "Chỉ chấp nhận file định dạng .xlsx")
            form_add_class.lift()

        # Hàm xử lý khi nhấn nút Lưu
        def save_data():
            class_name = class_name_var.get()
            subject_name = subject_name_var.get()
            session_count = session_count_var.get()
            file_path = file_path_var.get()


            device = "cuda:0" if torch.cuda.is_available() else "cpu"
            detection_net = load_retinaface("core/weights/mobilenet0.25_Final.pth", device)

            recognition_net = load_ghostfacenets()


            # Kiểm tra dữ liệu nhập
            if not all([class_name, file_path]):
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập đầy đủ thông tin!")
            elif not file_path.endswith(".xlsx"):
                messagebox.showerror("Lỗi", "Chỉ chấp nhận file định dạng .xlsx")
            else:
                # Xử lý lưu dữ liệu (giả sử bạn muốn in ra hoặc lưu vào DB)
                print(f"Tên lớp: {class_name}")
                print(f"Tên môn học: {subject_name}")
                print(f"Số buổi: {session_count}")
                print(f"Tệp Excel: {file_path}")
                #Create Class

                #create Student
                workbook =  load_workbook(file_path)
                sheet =  workbook.worksheets[0]
                for index,row in enumerate(sheet.iter_rows(values_only=True)):
                    if index ==0 :
                        continue
                    file_id = row[-1].split("id=")[-1]
                    download_url = f"https://drive.google.com/uc?id={file_id}&export=download"

                    print(download_url)
                    response = requests.get(download_url)

                    # Lưu ảnh vào file
                    if response.status_code == 200:
                        with open(f"static/images/{row[1]}.jpg", "wb") as f:
                            f.write(response.content)
                        print("Tải ảnh thành công!")

                    else:
                        messagebox.showerror("Lỗi", "Bạn phải chia sẻ công khai drive folder chứa hình ảnh. Mã lỗi:", response.status_code)
                        # return

                with app.app_context():
                    # Kiểm tra nếu lớp đã tồn tại
                    existing_class = db.session.query(Class).filter_by(name=class_name).first()
                    new_class = None
                    if existing_class:
                        print(f"Lớp '{class_name}' đã tồn tại.")
                        new_class = existing_class
                    else:
                    # Tạo lớp mới nếu không tồn tại
                        new_class = Class(
                            name=class_name,
                            created_at=datetime.datetime.now(),
                            updated_at=datetime.datetime.now()
                        )
                        db.session.add(new_class)
                        db.session.commit()
                        print(f"Lớp '{class_name}' đã được thêm thành công.")

                    for index,row in enumerate(sheet.iter_rows(values_only=True)):
                        if index ==0 :
                            continue
                        student_id = row[1]
                        student_name = row[2]
                        try:
                            profile_picture = f"static/images/{row[1]}.jpg"
                            profile_picture_vector = image_to_vector(profile_picture,detection_net,recognition_net,device)
                            check_student = db.session.query(Student).filter_by(id=student_id).first()
                            if check_student:
                                #update student
                                check_student.name = student_name
                                check_student.profile_picture = profile_picture
                                check_student.profile_picture_vector = profile_picture_vector
                                check_student.updated_at = datetime.datetime.now()
                                new_enrollment = Enrollment(student_id=student_id, class_id=new_class.id,
                                                            created_at=datetime.datetime.now(),
                                                            updated_at=datetime.datetime.now())
                                db.session.add(new_enrollment)
                                new_grade = Grade(student_id=student_id, class_id=new_class.id,
                                                  created_at=datetime.datetime.now(),
                                                  updated_at=datetime.datetime.now())
                                db.session.add(new_grade)
                                db.session.commit()

                            else:
                                #insert student
                                print("insert",student_id,student_name,profile_picture,profile_picture_vector)
                                new_student = Student(id=student_id,name=student_name,profile_picture=profile_picture,profile_picture_vector=profile_picture_vector,created_at=datetime.datetime.now(),updated_at=datetime.datetime.now())
                                db.session.add(new_student)
                                new_enrollment = Enrollment(student_id=student_id,class_id=new_class.id,created_at=datetime.datetime.now(),updated_at=datetime.datetime.now())
                                db.session.add(new_enrollment)
                                new_grade = Grade(student_id=student_id,class_id=new_class.id,created_at=datetime.datetime.now(),updated_at=datetime.datetime.now())
                                db.session.add(new_grade)
                                db.session.commit()
                        except Exception as e:

                            messagebox.showerror("Lỗi",
                                                 "Bạn phải chia sẻ công khai drive folder chứa hình ảnh. Mã lỗi:",
                                                 response.status_code)





                messagebox.showinfo("Thành công", "Dữ liệu đã được lưu!")
                form_add_class.destroy()


        # Thiết kế form
        widget_form_lop = ttk.LabelFrame(form_add_class, text="Thông tin lớp")
        widget_form_lop.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        widget_form_lop.rowconfigure(0, weight=1)
        widget_form_lop.columnconfigure(0, weight=1)
        widget_form_lop.columnconfigure(1, weight=5)
        widget_form_lop.columnconfigure(2, weight=1)

        label_ten_lop = tk.Label(widget_form_lop, text="Tên Lớp:")
        label_ten_lop.grid(row=0, column=0, padx=10, pady=5, sticky="e")
        entry_ten_lop = ttk.Entry(widget_form_lop, textvariable=class_name_var)
        entry_ten_lop.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        entry_ten_lop.insert(0, "Nhập tên lớp")
        entry_ten_lop.bind("<FocusIn>", lambda event: self.clear_entry(entry_ten_lop))

        # Time Picker

        label_tep_excel = tk.Label(widget_form_lop, text="Tệp Excel:")
        label_tep_excel.grid(row=1, column=0, padx=10, pady=5, sticky="e")
        entry_tep_excel = ttk.Entry(widget_form_lop, textvariable=file_path_var,  state="readonly")
        entry_tep_excel.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        button_chon_file = ttk.Button(widget_form_lop, text="Chọn File", command=select_file)
        button_chon_file.grid(row=1, column=2, padx=10, pady=5, sticky="e")



        ttk.Button(form_add_class, text="Lưu", style="Accent.TButton" , command=save_data, width=10).grid(row=2, column=0,
                                                                                                     columnspan=2,
                                                                                                     pady=0)






    def clear_entry(self,entry):
        entry.delete(0,tk.END)
