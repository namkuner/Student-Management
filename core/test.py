import json
import os

import numpy as np
import torch
import cv2
from detection.custom_config import cfg_mnet
from core.detection import detec
from recognition import rec
import torch.nn.functional as F

# device = "cuda:0" if torch.cuda.is_available() else "cpu"
# net = detec.load_retinaface(cfg_mnet, "weights/mobilenet0.25_Final.pth", device)



def euclidean_distance_squared(x, y):
    # Mở rộng chiều của x và y để có thể thực hiện phép trừ
    diff = x.unsqueeze(1) - y.unsqueeze(0)

    # Bình phương các phần tử và tính tổng theo chiều cuối cùng
    distances = torch.sum(diff ** 2, dim=-1)

    return distances
# students = db.session.query(Student).all()
#
# print(students)qqqq
def load(img_path):
    img = cv2.imread(img_path)
    h,w,_ =img.shape
    scale = h/1280
    if h > 1280 or w > 720:
        h = int(h/scale)
        w = int(w/scale)
        img = cv2.resize(img, (w,h), interpolation=cv2.INTER_AREA)
    print("h,w",h,w)
    priordata = detec.prior_box(cfg_mnet, h, w, device)


    res = detec.run(img,net,priordata,device)

    res = np.array(res)
    src = rec.run(res)
    return src

def run(img,net,priordata,device):
    res = detec.run(img, net, priordata, device)
    res = np.array(res)
    src = rec.run(res)
    return src

def compare_folder(folder_path):
    vector = []
    for i  in os.listdir(folder_path):
        img_path = os.path.join(folder_path,i)
        vector.append(load(img_path))

    vector = torch.stack(vector)
    res = euclidean_distance_squared(vector, vector)
    return  res

def emb_student(folder_path):
    sql_script = ""

    for i in os.listdir(folder_path):
        img_path = os.path.join(folder_path,i)
        vector = load(img_path)

        tensor_list = vector.tolist()
        tensor_json = json.dumps(tensor_list)

        mssv = i.split(".")[0]
        sql_script += f"UPDATE students SET profile_picture_vector = '{tensor_json}' WHERE id = '{mssv}';\n"



        # print("Sau khi giải")
        #
        # retrieved_json = tensor_json
        #
        #
        # tensor_list_restored = json.loads(retrieved_json)
        # tensor_restored = torch.tensor(tensor_list_restored)
        #
        # print(tensor_restored)

    return sql_script
if __name__ == "__main__":
    # img1 = load("image/te.jpg")

    from openpyxl import load_workbook
    import requests

    # Đường dẫn tới file Excel
    file_path = r"C:\Users\ADMIN.DESKTOP-VFGLH8R\Downloads\Testing.xlsx"

    # Tải workbook
    workbook = load_workbook(file_path)

    # Lấy sheet đầu tiên (theo thứ tự)
    first_sheet = workbook.worksheets[0]

    # Đọc dữ liệu từ sheet đầu tiên
    for row_index, row in enumerate(first_sheet.iter_rows(values_only=True)):
        if row_index == 0:
            continue  # Bỏ qua hàng đầu tiên (tiêu đề)

        if row and row[-1]:  # Kiểm tra nếu hàng không rỗng và cột cuối không rỗng

            file_id = row[-1].split("id=")[-1]
            download_url = f"https://drive.google.com/uc?id={file_id}&export=download"

            print(download_url)
            response = requests.get(download_url)

            # Lưu ảnh vào file
            if response.status_code == 200:
                with open(f"../static/images/{row[1]}.jpg", "wb") as f:
                    f.write(response.content)
                print("Tải ảnh thành công!")
                print(ngu)
            else:
                print("Tải ảnh thất bại. Mã lỗi:", response.status_code)


