

if __name__ == "__main__":
    import requests
    from openpyxl import load_workbook
    file_path = r"C:\Users\ADMIN.DESKTOP-VFGLH8R\Downloads\Testing.xlsx"
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[0]
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:
            continue
        file_id = row[-1].split("id=")[-1]
        download_url = f"https://drive.google.com/uc?id={file_id}&export=download"

        print(download_url)
        response = requests.get(download_url)

        # Lưu ảnh vào file
        if response.status_code == 200:
            with open(f"static/images/{row[1]}.jpg", "wb") as f:
                f.write(response.content)
            print("Tải ảnh thành công!")
        else:
            pass



